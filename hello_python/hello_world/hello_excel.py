import openpyxl

book = openpyxl.open("../assets/1.xlsx", read_only=False)
sheet = book.active

#all
for row in range(1, sheet.max_row+1):
    for col in range(0, sheet.max_column):
        print(sheet[row][col].value, end=" ")
    print()


#range from top left to bottom right
cells = sheet["B1" : "C2"]
for a, b in cells:
    print(a.value, b.value)


for a, b, c in sheet.iter_rows(min_row=1, max_row=2, min_col=0, max_col=3):
    print(a.value, b.value, c.value)
